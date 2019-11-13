# -*- coding: utf-8 -*-

import sys
import time
import numpy as np
import pandas as pd
import shutil
import openpyxl
import pathlib
import pywinauto
import os
from pywinauto.application import Application
from pywinauto.findwindows import WindowAmbiguousError, WindowNotFoundError
from pywinauto.controls.common_controls import TabControlWrapper
from pywinauto.keyboard import send_keys, KeySequenceError
import glob

from matlabconversionprog import matlab_conver_func,endurance_test
from threading import Thread, Lock
_db_lock = Lock() 

from concurrent import futures
import time
import functools
import multiprocessing  
thread_pool_executor = futures.ThreadPoolExecutor(max_workers=1)


def tk_after(target):
 
        @functools.wraps(target)
        def wrapper(self, *args, **kwargs):
            args = (self,) + args
            self.window.master.after(0, target, *args, **kwargs)
     
        return wrapper
 
 
def submit_to_pool_executor(executor):
    '''Decorates a method to be sumbited to the passed in executor'''
    def decorator(target):
 
        @functools.wraps(target)
        def wrapper(*args, **kwargs):
            result = executor.submit(target, *args, **kwargs)
            result.add_done_callback(executor_done_call_back)
            return result
 
        return wrapper
 
    return decorator
 
 
def executor_done_call_back(future):
    exception = future.exception()
    if exception:
        raise exception






class Analysis():
    def __init__(self,window,realrun):
        
        self.window=window
        self.realrun=realrun
        

    def analysis_threaded(self):
        #job_thread =Thread(target=self.analysis_alldays)
        #job_thread.start()   
        self.analysis_alldays()
    
    # Exit GUI cleanly
    
    def _quit(self):
        self.window.master.quit()
        self.window.master.destroy()
        


#    def analysis_russian(self):
#        
#        #list of all raw text files from all the days
#        lstrawtext=['raw_text_L1.txt','raw_text_L2.txt','raw_text_L3.txt']
#        #list of all the output files it needs to make
#        lst_treatg_input=[self.window.diff_tabs.sensor_input_set+"_AS_L1",self.window.diff_tabs.sensor_input_set+"_AS_L2",self.window.diff_tabs.sensor_input_set+"_AS_L3"]
#        lst_channels101_to_106=["101 (VDC)","201 (VDC)","102 (VDC)","202 (VDC)","103 (VDC)","203 (VDC)","104 (VDC)","204 (VDC)","105 (VDC)","205 (VDC)","106 (VDC)","206 (VDC)","107 (VDC)","207 (VDC)","108 (VDC)","208 (VDC)"]
#        lst_channels108_to_116=["109 (VDC)","209 (VDC)","110 (VDC)","210 (VDC)","111 (VDC)","211 (VDC)","112 (VDC)","212 (VDC)","113 (VDC)","213 (VDC)","114 (VDC)","214 (VDC)","115 (VDC)","215 (VDC)","116 (VDC)","216 (VDC)"]
#        #list of all the channels
#        lstchannels=lst_channels101_to_106+lst_channels108_to_116
#        
#        print("Preparing files for the TreatG software.. This takes a minute...\n")
#        #knowing the number of channels from the column number of 201
#        #Based on that the number of channels can be infered
#        #Now only these many number of channel files shall be created 
#        dfL1= pd.read_csv('raw_text_L1.txt')
#        location201=dfL1.columns.get_loc("201 (VDC)")
#        channels=int((location201-2)/2)
#        #Writing heading to all the files L1, L2,L3 with all the channel numbers so that 
#        #nextpart can append these files.The files will be L1_1 L1_2 till L1_channelnumber L2_channelnumber
#        #L3_channelnumber
#        for inputtreatg in lst_treatg_input:
#            for channelnum in range(0,channels):
#                with open(inputtreatg+"_"+str(channelnum+1), 'w') as f:
#                    f.write(' pos  uacc     Rts\n')
#        #This is the main part of the program
#                    
#         
#        for rawtext_day,inputtreatg in zip(lstrawtext,lst_treatg_input):
#            dfL1= pd.read_csv(rawtext_day)
#            #This for loop chooses the 4 rows, then the interchange among them happens in the desired order
#            #order from 1423 to 1234
#            for chose4rowspanda in range(0,33,4):
#                dfL1temp=dfL1.loc[chose4rowspanda:chose4rowspanda+3,:]
#                b,c,d=dfL1temp.iloc[1,:].copy(),dfL1temp.iloc[2,:].copy(),dfL1temp.iloc[3,:].copy()
#                dfL1temp.iloc[1,:],dfL1temp.iloc[2,:],dfL1temp.iloc[3,:]=c,d,b
#                #for every channel selected, we have to assign the correct column names
#                #It is like selecting a matrix using rows and columns channel number  gets the column name from the list
#                for channelnum in range(0,channels):
#                    with open(inputtreatg+"_"+str(channelnum+1), 'a') as f:
#                        for rows in range(0,4):
#                            f.write('  {0} {1:.6f}  {2:.4f}\n'.format(rows+1,dfL1temp[lstchannels[2*channelnum]].iloc[rows],dfL1temp[lstchannels[2*channelnum+1]].iloc[rows]))
#        print("Finished Preparing files for treatg...\n")
        
    def treatgsinglecycle_unused(self):
        app = Application().connect(title=u'D:\\Auto\\255678~1\\TreatG2.exe', class_name='ConsoleWindowClass')
        consolewindowclass = app.ConsoleWindowClass
        consolewindowclass.wait('ready')
        print("1")
        consolewindowclass.set_focus()
        #opening the popuplike box to show file names
        send_keys('{INS}')
        time.sleep(0.5)
        #selecting the first file
        send_keys('{RIGHT}{ENTER}')
        
        app.top_window().type_keys('{TAB}')
        
        time.sleep(0.5)
        app.top_window().type_keys('{TAB}')
        
        time.sleep(0.5)
        app.top_window().type_keys('{TAB}')
        
        time.sleep(0.5)
        send_keys('{DELETE}')
        time.sleep(0.5)
        send_keys("{2 down}"
                  "{2 up}"
                  "00.000"
                  ) # to type hello
        time.sleep(0.5)
        send_keys('{DOWN}')
        time.sleep(0.5)
        send_keys('{DELETE}')
        time.sleep(0.5)
        send_keys("{9 down}"
                  "{9 up}"
                  ".78335"
                  ) # to type hello
        
        
        send_keys('{ENTER}')
        time.sleep(1)
        send_keys('{ESC}')
        tempfile=pathlib.Path("D:/Auto/255678987/RESULT.DAT")
        required=pathlib.Path("D:/Auto/255678987/RESULT1.DAT")
        shutil.copy(tempfile,required)
        
        
        app.top_window().type_keys('{TAB}')
        time.sleep(0.5)
        
        send_keys('{INS}')
        time.sleep(0.5)
        
        send_keys('{RIGHT}{RIGHT}{ENTER}')
        
        app.top_window().type_keys('{TAB}')
        time.sleep(0.5)
        send_keys('{DOWN}')
        time.sleep(0.5)
        send_keys('{SPACE}')
        
        app.top_window().type_keys('{TAB}')
        send_keys('{ENTER}')
        time.sleep(1)
        send_keys('{ESC}')
        tempfile=pathlib.Path("D:/Auto/255678987/RESULT.DAT")
        required=pathlib.Path("D:/Auto/255678987/RESULT2.DAT")
        shutil.copy(tempfile,required)
        
        
        
        app.top_window().type_keys('{TAB}')
        
        time.sleep(0.5)
        app.top_window().type_keys('{TAB}')
        time.sleep(0.5)
        send_keys('{INS}')
        time.sleep(0.5)
        send_keys('{RIGHT}')
        time.sleep(0.5)
        send_keys('{RIGHT}')
        
        time.sleep(0.5)
        send_keys('{RIGHT}{ENTER}')
        
        app.top_window().type_keys('{TAB}')
        
        time.sleep(0.5)
        app.top_window().type_keys('{TAB}')
        send_keys('{ENTER}')
        time.sleep(1)
        send_keys('{ESC}')
        tempfile=pathlib.Path("D:/Auto/255678987/RESULT.DAT")
        required=pathlib.Path("D:/Auto/255678987/RESULT3.DAT")
        shutil.copy(tempfile,required)
        #app.Kill_()
    
    
    def checkfileexists(self, checkfilename):
        path_checkfilename=pathlib.Path(checkfilename)
        if not path_checkfilename.exists():
            print("Oops"+ checkfilename +" file doesn't exist!\n")
            self.window.master.destroy()
            sys.exit(0)
        else:
            print(checkfilename)
            return True


    def createdirectory(self,location):
        if not os.path.exists(location):
            try:
                os.mkdir(location)
            except OSError:
                print ("Creation of the directory locaton %s failed " % location)
                self.window.master.destroy()
                sys.exit(0)
            else:
                print ("Successfully created  directory %s " %location)
                
                
    
#    @submit_to_pool_executor(thread_pool_executor)
    def analysis_alldays(self):
        
        #INTERNAL when the file location can be got from the runs
        if self.window.diff_tabs.analysiscombo.get() == "Internal":
            dir_to_check=self.window.diff_tabs.base_folder
            sensorname=self.window.diff_tabs.sensor_input_set
            
        else:
            chosenfile=pathlib.Path(self.window.diff_tabs.defaultlocation)
            dir_to_check=chosenfile.parent
            
       
        #For checking if all the required files are present, if not the program exits
        for item in self.window.diff_tabs.ones_days:



            if self.window.diff_tabs.comboday_set == 'L1':
    
                self.checkfileexists(dir_to_check+'raw_text_L1.txt')
    
    
            elif self.window.diff_tabs.comboday_set == 'L2':
                self.checkfileexists(dir_to_check+'raw_text_L2.txt')
                
    
            elif self.window.diff_tabs.comboday_set == 'L3':
                self.checkfileexists(dir_to_check+'raw_text_L3.txt')
    
    
            elif self.window.diff_tabs.comboday_set == 'D4':
                self.checkfileexists(dir_to_check+'D4summary_temp1.txt')
                self.checkfileexists(dir_to_check+'D4summary_temp2.txt')
                self.checkfileexists(dir_to_check+'D4summary_temp3.txt')
                self.checkfileexists(dir_to_check+'raw_text_D4_1_33.txt')
                self.checkfileexists(dir_to_check+'raw_text_D4_2_33.txt')
                self.checkfileexists(dir_to_check+'raw_text_D4_3_33.txt')
                
            elif (self.window.diff_tabs.comboday_set == 'D5' ):
                self.checkfileexists(dir_to_check+'D5summary_temp1.txt')
                self.checkfileexists(dir_to_check+'D5summary_temp2.txt')
                self.checkfileexists(dir_to_check+'D5summary_temp3.txt')
                
                    
            elif (self.window.diff_tabs.comboday_set == 'D6' ):
                self.checkfileexists(dir_to_check+'raw_text_D6.txt')
            else:
                print("\n")

            
            
        #concatenating all the names of the days that need to be analyzed    
        if len(self.window.diff_tabs.ones_days)>0:
           seperator = '_'
           dir_name_analysis="analysis"+seperator.join(self.window.diff_tabs.ones_days)
        else:
           dir_name_analysis="complete_dummy"
           
        
        
        
        
        #Create analysis folder where analysis is going to happen
        self.dir_where_analy_happ=pathlib.Path(dir_to_check+dir_name_analysis)
        self.createdirectory(str(self.dir_where_analy_happ))
        
        
        #copy files that are  present from  dir_to_check to the dir_where_analy_happ
        #Basically, these are really data files
        for comboday in self.window.diff_tabs.ones_days:
            if comboday=="L1":
                reqfiles=['raw_text_L1.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(dir_to_check.joinpath(file))
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
            elif comboday=="L2":
                reqfiles=['raw_text_L2.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(dir_to_check.joinpath(file))
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
            elif  comboday=="L3":
                reqfiles=['raw_text_L3.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(dir_to_check.joinpath(file))
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
                
            elif  comboday=="D4":
                reqfiles=['D4summary_temp1.txt','D4summary_temp2.txt','D4summary_temp3.txt','raw_text_D4_1_33.txt','raw_text_D4_2_33.txt','raw_text_D4_3_33.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(dir_to_check.joinpath(file))
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
            elif comboday=="D5":
                reqfiles=['D5summary_temp1.txt','D5summary_temp2.txt','D5summary_temp3.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(dir_to_check.joinpath(file))
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
                
            elif comboday=="D6":
                reqfiles=['raw_text_D6.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(dir_to_check.joinpath(file))
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
           
            else:
                print("\n")
        
        
        
        
        
        
        
        
        #copy files that are not present from resources folder to the dir_where_analy_happ
        for comboday in self.window.diff_tabs.zero_days:
            if comboday=="L1":
                reqfiles=['raw_text_L1.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+file)
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
            elif comboday=="L2":
                reqfiles=['raw_text_L2.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+file)
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
            elif  comboday=="L3":
                reqfiles=['raw_text_L3.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+file)
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
                
            elif  comboday=="D4":
                reqfiles=['D4summary_temp1.txt','D4summary_temp2.txt','D4summary_temp3.txt','raw_text_D4_1_33.txt','raw_text_D4_2_33.txt','raw_text_D4_3_33.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+file)
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
            elif comboday=="D5":
                reqfiles=['D5summary_temp1.txt','D5summary_temp2.txt','D5summary_temp3.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+file)
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
                
            elif comboday=="D6":
                reqfiles=['raw_text_D6.txt']
                for file in reqfiles:
                    from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+file)
                    to_file=pathlib.Path(self.dir_where_analy_happ.joinpath(file))
                    shutil.copy(from_file,to_file)
           
            else:
                print("\n")
        print("Copying the missing files from resource folder finished...\n") 
        
        
        self.Dividing_into_channelsL123() 
        self.placing_each_channel_direc()
##            
    
    def placing_each_channel_direc(self):
        
        lst_channels101_to_106=["101 (VDC)","201 (VDC)","102 (VDC)","202 (VDC)","103 (VDC)","203 (VDC)","104 (VDC)","204 (VDC)","105 (VDC)","205 (VDC)","106 (VDC)","206 (VDC)","107 (VDC)","207 (VDC)","108 (VDC)","208 (VDC)"]
        lst_channels108_to_116=["109 (VDC)","209 (VDC)","110 (VDC)","210 (VDC)","111 (VDC)","211 (VDC)","112 (VDC)","212 (VDC)","113 (VDC)","213 (VDC)","114 (VDC)","214 (VDC)","115 (VDC)","215 (VDC)","116 (VDC)","216 (VDC)"]
        #list of all the channels
        lstchannels=lst_channels101_to_106+lst_channels108_to_116
        lst_all101channels=lstchannels[0::2]
        
        
         #list of all raw text files from all the days
        lstD4temp_33files=['raw_text_D4_1_33.txt','raw_text_D4_2_33.txt','raw_text_D4_3_33.txt']
        lst_of_excelfiles=['(AS)_adj_param_m40.xlsm','(AS)_adj_param_p70.xlsm','(AS)_adj_param_p20.xlsm']
        lst_ten_numbers=list(range(0,10))
        lstfinalD45files=['finalD45_temp1.txt','finalD45_temp2.txt','finalD45_temp3.txt']
        lst_D4summary=['D4summary_temp1.txt','D4summary_temp2.txt','D4summary_temp3.txt']
        lst_D5summary=['D5summary_temp1.txt','D5summary_temp2.txt','D5summary_temp3.txt']
        
        print("Starting Placing each channel in directory \n")
        
        
        for channelnum,channelname in enumerate(self.window.diff_tabs.channelnames):
            
            channelpath=pathlib.Path(self.dir_where_analy_happ.joinpath(channelname))
            self.createdirectory(str(channelpath))
            
            treatgchannelpath=pathlib.Path(channelpath.joinpath("treatg"))
            
            self.createdirectory(str(treatgchannelpath))
            pattern="*L[1-3]_"+str(channelnum+1)+"*"
            #for file in glob.glob(str(pathlib.Path(self.dir_where_analy_happ).joinpath("*L[1-3]_1*"))):
           
            for file in glob.glob(str(pathlib.Path(self.dir_where_analy_happ).joinpath(pattern))):
                shutil.copy(file,treatgchannelpath)
            
            
            print(f"Finished Placing channelnum {channelnum} in directory \n")
            #Now all the required files are present for treatg analysis folder
            
            #self.treatgsinglecycle(treatgchannelpath)
            print(f"Satarting Treatg Analysis channelnum  {channelnum}\n")
            #simulating the situation of running the treatg software
            from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+"RESULT1.DAT")
            shutil.copy(from_file,channelpath)    
            from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+"RESULT2.DAT")
            shutil.copy(from_file,channelpath)   
            from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+"RESULT3.DAT")
            shutil.copy(from_file,channelpath) 
            
            
            #copy the excel files to the channel location 
            for excelfile in lst_of_excelfiles:
                from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+excelfile)
                shutil.copy(from_file,channelpath)    
            
            
            
        
            print(f"Finished copying excel files\n")
            #writing _33 information of the  channel into excel files
            dir_for_analy=channelpath
            #For all the files in the list writing 33 values in to the desired excel cells
            for  excelfile, file33 in zip(lst_of_excelfiles,lstD4temp_33files):
                            
                df_temp_33=pd.read_csv(dir_for_analy.parent.joinpath(file33))
                df_temp_33=df_temp_33.filter(like=lst_all101channels[channelnum])
                for item in lst_ten_numbers:
                    
                    
                    srcfile = openpyxl.load_workbook(dir_for_analy.joinpath(excelfile),read_only=False, keep_vba= True)#to open the excel sheet and if it has macros
                    sheetname=srcfile["sheetsai"]
                    sheetname.cell(row=item+4,column=3).value = df_temp_33.loc[item].values[0]#write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
                    name_file=excelfile.split(".")[0]+"_modified"+".xlsm"
                srcfile.save(dir_for_analy.joinpath(name_file))#save it as a new file, the original file is untouched and here I am saving it as xlsm(m here denotes macros).
                 
            lst_of_xlsmfiles_modified=['(AS)_adj_param_m40_modified.xlsm','(AS)_adj_param_p70_modified.xlsm','(AS)_adj_param_p20_modified.xlsm']   
            lstsix_numbers=list(range(0,6))
            
            
            
            #creating finalD45_temp1 2 and 3  files which contains  the desired six rows by six colummns that can be used in the program 
            for finalfile,D4summ_file,D5summ_file in zip(lstfinalD45files,lst_D4summary,lst_D5summary):
                df_D4summ=pd.read_csv(dir_for_analy.parent.joinpath(D4summ_file))
                df_D5summ=pd.read_csv(dir_for_analy.parent.joinpath(D5summ_file))
                df_D4_for_channel=df_D4summ.filter(like=lst_all101channels[channelnum])
                df_D5_for_channel=df_D5summ.filter(like=lst_all101channels[channelnum])
                df_D45_for_channel=pd.concat([df_D4_for_channel,df_D5_for_channel],axis=1)
                df_D45_for_channel.to_csv(dir_for_analy.joinpath(finalfile),index=False)
            
            
            #For all the files in the list writing D45 values in to the desired excel cells
            for  xlsmfile,fileD45 in zip(lst_of_xlsmfiles_modified,lstfinalD45files):
                
                df_temp_45=pd.read_csv(dir_for_analy.joinpath(fileD45))
                df_temp_45=df_temp_45.filter(like=lst_all101channels[channelnum])
                for rows in lstsix_numbers:
                    for columns in lstsix_numbers:
                    
                        
                        
                        
                        srcfile = openpyxl.load_workbook(dir_for_analy.joinpath(xlsmfile),read_only=False, keep_vba= True)#to open the excel sheet and if it has macros
                        sheetname=srcfile["sheetsai"]
                        sheetname.cell(row=rows+19,column=columns+2).value = df_temp_45.iloc[rows,columns] #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
                srcfile.save(dir_for_analy.joinpath(channelname+xlsmfile))
                
            #Now that the desired excel files are written deleting the extra unnecessary files 
            
            for file1,file2 in zip(lst_of_excelfiles,lst_of_xlsmfiles_modified):
                file_to_rem = dir_for_analy.joinpath(file1)
                file_to_rem.unlink()
                file_to_rem = dir_for_analy.joinpath(file2)
                file_to_rem.unlink()
            
            #Makng the desired array as the input for matlab conversion program
            #df_for_matlab=df_temp_33=pd.read_csv(dir_for_analy.joinpath("raw_text_D4_1_33.txt"),header=None)
            #I think temp3 d4 first two columns
            df_for_matlab=pd.read_csv(dir_for_analy.joinpath('finalD45_temp3.txt'))
            desired_matrix=df_for_matlab.iloc[:,0:2].values
            
            #copying sample1.csv and sample2.csv into  respective channel folders
            #Also copied the endurance files also
            reqfiles=["sample1.csv","sample2.csv","endu_sample1.csv","endu_sample2.csv","endu_sample3.csv"]
            for file in reqfiles:
                from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+file)
                shutil.copy(from_file,dir_for_analy)
            
            
            
           
            
            #Calling the matlab conversion program
            matlab_conver_func(desired_matrix,dir_for_analy,channelname)
            
        
            #Calling the endurance test program
            df_endu=pd.read_csv(dir_for_analy.parent.joinpath('raw_text_D6.txt'))
            df_endu_channels=df_endu[[ lstchannels[channelnum*2],lstchannels[channelnum*2+1]  ]]
            endurance_test(df_endu_channels,dir_for_analy,channelname)
            
            
            
            
            
            
            
            
            
            
            
            
            
            
            
#            dfD4summary_temp1=pd.read_csv(self.dir_where_analy_happ.joinpath('D4summary_temp1.txt'))
#            
#            #dfD4summary_temp1.filter(like="101 (VDC)")
#            dfD4summary_temp1.filter(like=lst_all101channels[channelnum])
            
            



    def finalanalysisD45(self):
        dfD4summary_temp1=pd.read_csv('D4summary_temp1.txt')
        dfD5_temp1_pos5= pd.read_csv('raw_text_D5_1_5.txt')
        dfD5_temp1_pos6= pd.read_csv('raw_text_D5_1_6.txt')

        df_finalD45_temp1=pd.concat([dfD4summary_temp1,dfD5_temp1_pos5,dfD5_temp1_pos6],axis=1)
        df_finalD45_temp1.to_csv('finalD45_temp1.txt',index=False)


        dfD4summary_temp2=pd.read_csv('D4summary_temp2.txt')
        dfD5_temp2_pos5= pd.read_csv('raw_text_D5_2_5.txt')
        dfD5_temp2_pos6= pd.read_csv('raw_text_D5_2_6.txt')

        df_finalD45_temp2=pd.concat([dfD4summary_temp2,dfD5_temp2_pos5,dfD5_temp2_pos6],axis=1)
        df_finalD45_temp2.to_csv('finalD45_temp2.txt',index=False)

        dfD4summary_temp3=pd.read_csv('D4summary_temp3.txt')
        dfD5_temp3_pos5= pd.read_csv('raw_text_D5_3_5.txt')
        dfD5_temp3_pos6= pd.read_csv('raw_text_D5_3_6.txt')

        df_finalD45_temp3=pd.concat([dfD4summary_temp3,dfD5_temp3_pos5,dfD5_temp3_pos6],axis=1)
        df_finalD45_temp3.to_csv('finalD45_temp3.txt',index=False)
        
        
        dfD4_temp1_pos33=pd.read_csv('raw_text_D4_1_33.txt')
        dfD4_temp2_pos33=pd.read_csv('raw_text_D4_2_33.txt')
        dfD4_temp3_pos33=pd.read_csv('raw_text_D4_3_33.txt')



        
    def treatgsinglecycle(self,treatgchannelpath):
        
        from_file=pathlib.Path(self.window.diff_tabs.resourcelocation+"TreatG2.exe")
        to_file=pathlib.Path(treatgchannelpath)
        shutil.copy(from_file,to_file)
        treatg_exe_place=pathlib.Path(to_file.joinpath("TreatG2.exe"))
        
        os.startfile(str(treatg_exe_place))
        
        old_path=pathlib.PureWindowsPath(treatg_exe_place)
        new_path=list(old_path.parts)
        var=[]
        for item in new_path:
            if len(item)>6:
                item=item[0:6]+"~"+"1"
                print(item)
                var.append(item)
        
    
            else:
                print(item)
                var.append(item)

        new_path="\\".join(var)+"\\TreatG2.exe"
        new_path=pathlib.PureWindowsPath(new_path)
        new_path=u"{}".format(str(new_path))
        
        
        app = Application().connect(title=new_path, class_name='ConsoleWindowClass')
        consolewindowclass = app.ConsoleWindowClass
        consolewindowclass.wait('ready')
        print("1")
        consolewindowclass.set_focus()
        #opening the popuplike box to show file names
        send_keys('{INS}')
        time.sleep(0.5)
        #selecting the first file
        send_keys('{RIGHT}{ENTER}')
        
        app.top_window().type_keys('{TAB}')
        
        time.sleep(0.5)
        app.top_window().type_keys('{TAB}')
        
        time.sleep(0.5)
        app.top_window().type_keys('{TAB}')
        
        time.sleep(0.5)
        send_keys('{DELETE}')
        time.sleep(0.5)
        send_keys("{2 down}"
                  "{2 up}"
                  "00.000"
                  ) # to type hello
        time.sleep(0.5)
        send_keys('{DOWN}')
        time.sleep(0.5)
        send_keys('{DELETE}')
        time.sleep(0.5)
        send_keys("{9 down}"
                  "{9 up}"
                  ".78335"
                  ) # to type hello
        
        
        send_keys('{ENTER}')
        time.sleep(1)
        send_keys('{ESC}')
        
        #to file is the treatgchannel path which is inside analysis_l1_L2+channel+treatgchannel
        tempfile=pathlib.Path(to_file.joinpath("RESULT.DAT"))
        result1path=pathlib.Path(to_file.joinpath("RESULT1.DAT"))
        #tempfile=pathlib.Path("D:/Auto/255678987/RESULT.DAT")
        #required=pathlib.Path("D:/Auto/255678987/RESULT1.DAT")
        shutil.copy(tempfile,result1path)
        
        
        app.top_window().type_keys('{TAB}')
        time.sleep(0.5)
        
        send_keys('{INS}')
        time.sleep(0.5)
        
        send_keys('{RIGHT}{RIGHT}{ENTER}')
        
        app.top_window().type_keys('{TAB}')
        time.sleep(0.5)
        send_keys('{DOWN}')
        time.sleep(0.5)
        send_keys('{SPACE}')
        
        app.top_window().type_keys('{TAB}')
        send_keys('{ENTER}')
        time.sleep(1)
        send_keys('{ESC}')
        
        tempfile=pathlib.Path(to_file.joinpath("RESULT.DAT"))
        result2path=pathlib.Path(to_file.joinpath("RESULT2.DAT"))
        #tempfile=pathlib.Path("D:/Auto/255678987/RESULT.DAT")
        #required=pathlib.Path("D:/Auto/255678987/RESULT2.DAT")
        shutil.copy(tempfile,result2path)
        
        
        
        app.top_window().type_keys('{TAB}')
        
        time.sleep(0.5)
        app.top_window().type_keys('{TAB}')
        time.sleep(0.5)
        send_keys('{INS}')
        time.sleep(0.5)
        send_keys('{RIGHT}')
        time.sleep(0.5)
        send_keys('{RIGHT}')
        
        time.sleep(0.5)
        send_keys('{RIGHT}{ENTER}')
        
        app.top_window().type_keys('{TAB}')
        
        time.sleep(0.5)
        app.top_window().type_keys('{TAB}')
        send_keys('{ENTER}')
        time.sleep(1)
        send_keys('{ESC}')
        tempfile=pathlib.Path(to_file.joinpath("RESULT.DAT"))
        result3path=pathlib.Path(to_file.joinpath("RESULT3.DAT"))
        shutil.copy(tempfile,result3path)
#        tempfile=pathlib.Path("D:/Auto/255678987/RESULT.DAT")
#        required=pathlib.Path("D:/Auto/255678987/RESULT3.DAT")
        
        #Now copy all the three files to the previous folder which is channelfolder
        shutil.copy(result1path,to_file.parent)
        shutil.copy(result2path,to_file.parent)
        shutil.copy(result3path,to_file.parent)
        
        
        app.Kill_()
        

    
        
            
    def Dividing_into_channelsD45(self):
        
        
        
        
        #list of all raw text files from all the days
        lstrawtextwithoutpath=['D4summary_temp1.txt','D4summary_temp2.txt','D4summary_temp3.txt','D5summary_temp1.txt','D5summary_temp2.txt','D5summary_temp3.txt']
        lstrawtext=list(map(lambda x:str(self.dir_where_analy_happ.joinpath(x)),lstrawtextwithoutpath))
        
        #list of all the output files it needs to make
        
        
        
        d4input_withoutpath=[self.window.diff_tabs.sensor_input_set+"D4summary_temp1","1_"+self.window.diff_tabs.sensor_input_set+"D4summary_temp2","1_"+self.window.diff_tabs.sensor_input_set+"D4summary_temp3"]
        lst_treatg_input=list(map(lambda x:str(self.dir_where_analy_happ.joinpath(x)),lst_treatg_input_withoutpath))
        
        
        
        lst_channels101_to_106=["101 (VDC)","201 (VDC)","102 (VDC)","202 (VDC)","103 (VDC)","203 (VDC)","104 (VDC)","204 (VDC)","105 (VDC)","205 (VDC)","106 (VDC)","206 (VDC)","107 (VDC)","207 (VDC)","108 (VDC)","208 (VDC)"]
        lst_channels108_to_116=["109 (VDC)","209 (VDC)","110 (VDC)","210 (VDC)","111 (VDC)","211 (VDC)","112 (VDC)","212 (VDC)","113 (VDC)","213 (VDC)","114 (VDC)","214 (VDC)","115 (VDC)","215 (VDC)","116 (VDC)","216 (VDC)"]
        #list of all the channels
        lstchannels=lst_channels101_to_106+lst_channels108_to_116
        lst_all101channels=lstchannels[0::2]
        
        D4summary_temp1.filter(like="101 (VDC)")
        
        
        
        print("Preparing files for the TreatG software.. D45...This takes a minute...\n")
        
        
        channels=self.window.diff_tabs.no_channels
        #Writing heading to all the files L1, L2,L3 with all the channel numbers so that 
        #nextpart can append these files.The files will be L1_1 L1_2 till L1_channelnumber L2_channelnumber
        #L3_channelnumber
        for inputtreatg in lst_treatg_input:
            for channelnum in range(0,channels):
                with open(inputtreatg+"_"+str(channelnum+1), 'w') as f:
                    f.write(' pos  uacc     Rts\n')
        #This is the main part of the program
                    
         
        for rawtext_day,inputtreatg in zip(lstrawtext,lst_treatg_input):
            dfL1= pd.read_csv(rawtext_day)
            #This for loop chooses the 4 rows, then the interchange among them happens in the desired order
            #order from 1423 to 1234
            for chose4rowspanda in range(0,((self.window.diff_tabs.tempruns_set*4)-3),4):
                dfL1temp=dfL1.loc[chose4rowspanda:chose4rowspanda+3,:]
                b,c,d=dfL1temp.iloc[1,:].copy(),dfL1temp.iloc[2,:].copy(),dfL1temp.iloc[3,:].copy()
                dfL1temp.iloc[1,:],dfL1temp.iloc[2,:],dfL1temp.iloc[3,:]=c,d,b
                #for every channel selected, we have to assign the correct column names
                #It is like selecting a matrix using rows and columns channel number  gets the column name from the list
                for channelnum in range(0,channels):
                    with open(inputtreatg+"_"+str(channelnum+1)+".DAT", 'a') as f:
                        for rows in range(0,4):
                            f.write('  {0} {1:.6f}  {2:.4f}\n'.format(rows+1,dfL1temp[lstchannels[2*channelnum]].iloc[rows],dfL1temp[lstchannels[2*channelnum+1]].iloc[rows]))
        
        
        
        print("Finished Preparing files for treatg L123...\n")   
        




    
        
            
    def Dividing_into_channelsL123(self):
        
        
        #list of all raw text files from all the days
        lstrawtextwithoutpath=['raw_text_L1.txt','raw_text_L2.txt','raw_text_L3.txt']
        lstrawtext=list(map(lambda x:str(self.dir_where_analy_happ.joinpath(x)),lstrawtextwithoutpath))
        
        #list of all the output files it needs to make
        
        
        
        lst_treatg_input_withoutpath=["1_"+self.window.diff_tabs.sensor_input_set+"_AS_L1","1_"+self.window.diff_tabs.sensor_input_set+"_AS_L2","1_"+self.window.diff_tabs.sensor_input_set+"_AS_L3"]
        lst_treatg_input=list(map(lambda x:str(self.dir_where_analy_happ.joinpath(x)),lst_treatg_input_withoutpath))
        
        
        
        lst_channels101_to_106=["101 (VDC)","201 (VDC)","102 (VDC)","202 (VDC)","103 (VDC)","203 (VDC)","104 (VDC)","204 (VDC)","105 (VDC)","205 (VDC)","106 (VDC)","206 (VDC)","107 (VDC)","207 (VDC)","108 (VDC)","208 (VDC)"]
        lst_channels108_to_116=["109 (VDC)","209 (VDC)","110 (VDC)","210 (VDC)","111 (VDC)","211 (VDC)","112 (VDC)","212 (VDC)","113 (VDC)","213 (VDC)","114 (VDC)","214 (VDC)","115 (VDC)","215 (VDC)","116 (VDC)","216 (VDC)"]
        #list of all the channels
        lstchannels=lst_channels101_to_106+lst_channels108_to_116
        
        
        print(f"The analysis happens in {self.dir_where_analy_happ}")
        for i,j in zip(lstrawtext,lst_treatg_input):
            print(i,j)
        
        print("Preparing files for the TreatG software.. L123...This takes upto 5  minutes...\n")
        #knowing the number of channels from the column number of 201
        #Based on that the number of channels can be infered
        #Now only these many number of channel files shall be created 
#        dfL1= pd.read_csv('raw_text_L1.txt')
#        location201=dfL1.columns.get_loc("201 (VDC)")
#        channels=int((location201-2)/2)
        
        channels=self.window.diff_tabs.no_channels
        #Writing heading to all the files L1, L2,L3 with all the channel numbers so that 
        #nextpart can append these files.The files will be L1_1 L1_2 till L1_channelnumber L2_channelnumber
        #L3_channelnumber
        for inputtreatg in lst_treatg_input:
            for channelnum in range(0,channels):
                with open(inputtreatg+"_"+str(channelnum+1)+".DAT", 'w') as f:
                    f.write(' pos  uacc     Rts\n')
        #This is the main part of the program
                    
         
        for rawtext_day,inputtreatg in zip(lstrawtext,lst_treatg_input):
            dfL1= pd.read_csv(rawtext_day)
            #This for loop chooses the 4 rows, then the interchange among them happens in the desired order
            #order from 1423 to 1234
            for chose4rowspanda in range(0,((self.window.diff_tabs.tempruns_set*4)-3),4):
                dfL1temp=dfL1.loc[chose4rowspanda:chose4rowspanda+3,:]
                b,c,d=dfL1temp.iloc[1,:].copy(),dfL1temp.iloc[2,:].copy(),dfL1temp.iloc[3,:].copy()
                dfL1temp.iloc[1,:],dfL1temp.iloc[2,:],dfL1temp.iloc[3,:]=c,d,b
                #for every channel selected, we have to assign the correct column names
                #It is like selecting a matrix using rows and columns channel number  gets the column name from the list
                for channelnum in range(0,channels):
                    with open(inputtreatg+"_"+str(channelnum+1)+".DAT", 'a') as f:
                        for rows in range(0,4):
                            f.write('  {0} {1:.6f}  {2:.4f}\n'.format(rows+1,dfL1temp[lstchannels[2*channelnum]].iloc[rows],dfL1temp[lstchannels[2*channelnum+1]].iloc[rows]))
        
        
        
        print("Finished Preparing files for treatg L123...\n")   
        
        
    def preparing_reports(self,dir_for_analy):
        
#        lstresult_files=['RESULT1.DAT','RESULT2.DAT','RESULT3.DAT']
#        lstfinalD45files=['finalD45_temp1.txt','finalD45_temp2.txt','finalD45_temp3.txt']        
        lstD4temp_33files=['raw_text_D4_1_33.txt','raw_text_D4_2_33.txt','raw_text_D4_3_33.txt']
#        listof_filesrequired=lstresult_files+lstfinalD45files+lstD4temp_33files
#        #printing all the required file names 
#        for item in listof_filesrequired:
#            print(item) 
#            
#        
#          
#        #For checking if all the required files are present, if not the program exits
#        for item in listof_filesrequired:
#            self.checkfileexists(dir_for_analy.joinpath(item))
#        
        
            
        lst_of_excelfiles=['(AS)_adj_param_m40.xlsx','(AS)_adj_param_p70.xlsx','(AS)_adj_param_p20.xlsx']
        
        
        #g_sheet=srcfile.sheetnames
        
        #sheetname = srcfile.get_sheet_by_name('sheetsai')#get sheetname from the file
        #sheetname['C4']= 55.568 #write something in B2 cell of the supplied sheet
        
        lst_ten_numbers=list(range(0,10))
        
        
        #For all the files in the list writing 33 values in to the desired excel cells
        for  excelfile, file33 in zip(lst_of_excelfiles,lstD4temp_33files):
            for num in lst_ten_numbers:
                
                
                df_temp_33=pd.read_csv(dir_for_analy.joinpath(file33),header=None)
                
                srcfile = openpyxl.load_workbook(dir_for_analy.joinpath(excelfile),read_only=False, keep_vba= True)#to open the excel sheet and if it has macros
                sheetname=srcfile["sheetsai"]
                sheetname.cell(row=item+4,column=3).value = df_temp_33.loc[item,0] #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
            srcfile.save(dir_for_analy.joinpath(excelfile.split(".")[0]+"_modified"+".xlsm"))#save it as a new file, the original file is untouched and here I am saving it as xlsm(m here denotes macros).
             
        lst_of_xlsmfiles_modified=['(AS)_adj_param_m40_modified.xlsm','(AS)_adj_param_p70_modified.xlsm','(AS)_adj_param_p20_modified.xlsm']   
        lstsix_numbers=list(range(0,6))
        
        #For all the files in the list writing D45 values in to the desired excel cells
        for  xlsmfile,fileD45 in zip(lst_of_xlsmfiles_modified,lstD4temp_33files):
            for rows in lstsix_numbers:
                for columns in lstsix_numbers:
                
                    
                    df_temp_45=pd.read_csv(dir_for_analy.joinpath(fileD45),header=None)
                    
                    srcfile = openpyxl.load_workbook(dir_for_analy.joinpath(xlsmfile),read_only=False, keep_vba= True)#to open the excel sheet and if it has macros
                    sheetname=srcfile["sheetsai"]
                    sheetname.cell(row=rows+19,column=columns+2).value = df_temp_45.loc[rows,columns] #write to row 1,col 1 explicitly, this type of writing is useful to write something in loops
            srcfile.save(dir_for_analy.joinpath(sensorname+xlsmfile))
            
        #Now that the desired excel files are written deleting the extra unnecessary files 
        
        for file1,file2 in zip(lst_of_excelfiles,lst_of_xlsmfiles_modified):
            file_to_rem = dir_for_analy.joinpath(file1)
            file_to_rem.unlink()
            file_to_rem = dir_for_analy.joinpath(file2)
            file_to_rem.unlink()
        
        #Makng the desired array as the input for matlab conversion program
        df_for_matlab=df_temp_33=pd.read_csv(dir_for_analy.joinpath("raw_text_D4_1_33.txt"),header=None)
        desired_matrix=df_for_matlab.iloc[:,0:2].values
        
        
        #Calling the matlab conversion program
        matlab_conver_func(desired_matrix,dir_for_analy,sensorname)
        
    
    
    
    
        
        
                
    